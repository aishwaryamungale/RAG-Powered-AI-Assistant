import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from config.settings import settings

class FeedbackService:
    def __init__(self):
        self.feedback_file = settings.FEEDBACK_FILE
    
    def save_feedback(self, query, advice, feedback_type, rating):
        if not query or not advice:
            return "Missing query or advice."
        
        if os.path.exists(self.feedback_file):
            wb = load_workbook(self.feedback_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Query", "Advice", "Feedback", "Rating", "Timestamp"])
        
        ws.append([
            query[:500],  # Truncate long queries
            advice[:1000],  # Truncate long advice
            feedback_type,
            rating,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ])
        
        wb.save(self.feedback_file)
        return "Feedback saved successfully!"